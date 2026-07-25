import sys
import os
import json

from openpyxl import Workbook
from local_db import get_connection



script_dir = os.path.dirname(__file__)
posettings = os.path.join(script_dir, 'posettings.json')
with open(posettings, 'r') as json_file:
    data = json.load(json_file)



def generate_excel(df,start_date):
    po_file_path = data.get('POFilePath')
    Upload_folder = os.path.join(po_file_path,"PO_Created_Files")
    os.makedirs(Upload_folder, exist_ok=True)
    container = [1,1,1,1]

    wb = Workbook()
    ws1 = wb.create_sheet("HAR")
    ws1.append(["Sr No", "PO No","Address", "Code", "SKU","QTY"])
    ws2 = wb.create_sheet("BAN")
    ws2.append(["Sr No", "PO No","Address", "Code", "SKU","QTY"])
    ws3 = wb.create_sheet("MUM")
    ws3.append(["Sr No", "PO No","Address", "Code", "SKU","QTY"])
    
    
    default_sheet = wb.get_sheet_by_name('Sheet')
    if default_sheet:
        wb.remove(default_sheet) 
    for index, row in df.iterrows():
        GSTN = row['GSTN']
        po_number = row['PO_Number']
        Final_Address = row['Final_Address']
        Coad = row['Coad']
        mapped_SKUs = row['Mapped_SKUs']
        quantities = row['Quantities']
        company = row['Company']
        today_date = row['Date']

        
        if GSTN == "06":
            for i, (coad, sku, qty) in enumerate(zip(Coad, mapped_SKUs, quantities), start=1):
                if i == 1:
                    row_data = [container[0], po_number, Final_Address, coad, sku, qty]
                    ws1.append(row_data)
                else:
                    row_data2 = [None, None,  None, coad, sku, qty]
                    ws1.append(row_data2)
            ws1.append([None] * 7)  # Append an empty row
            container[0] += 1
        
        elif GSTN == "29":
            for i, (coad, sku, qty) in enumerate(zip(Coad, mapped_SKUs, quantities), start=1):
                if i == 1:
                    row_data = [container[1], po_number, Final_Address, coad, sku, qty]
                    ws2.append(row_data)
                else:
                    row_data2 = [None, None, None,  coad, sku, qty]
                    ws2.append(row_data2)
            ws2.append([None] * 7)  # Append an empty row
            container[1] += 1
        
        elif GSTN == "27":
            for i, (coad, sku, qty) in enumerate(zip(Coad, mapped_SKUs, quantities), start=1):
                if i == 1:
                    row_data = [container[2], po_number,Final_Address, coad, sku, qty]
                    ws3.append(row_data)
                else:
                    row_data2 = [None, None, None, coad, sku, qty]
                    ws3.append(row_data2)
            ws3.append([None] * 7)  # Append an empty row
            container[2] += 1
    
    excel_file_path = os.path.join(Upload_folder, f'{company}_PO_File_{start_date}_to_{today_date}.xlsx')
    wb.save(excel_file_path)
    print(f"Excel file saved at: {excel_file_path}")     


# -------------------------------------------------------
# CHECK PO EXISTS
# -------------------------------------------------------
def check_data_in_database(ponumber):
   conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT ponumber FROM compiledpurchaseorder WHERE ponumber = ?;",
        (ponumber,)
    )

    existing_pos = cur.fetchall()

    cur.close()
    conn.close()

    if existing_pos:
        print("PO Exist = ", existing_pos)
        return True
    else:
        return False


# -------------------------------------------------------
# UPLOAD / UPDATE PO
# -------------------------------------------------------
def upload_po(df):
   conn = get_connection()
    cur = conn.cursor()

    for index, row in df.iterrows():
        GSTN = row['GSTN']
        po_number = row['PO_Number']
        Final_Address = row['Final_Address']
        Coad = row['Coad']
        mapped_SKUs = row['Mapped_SKUs']
        quantities = row['Quantities']
        company = row['Company']
        start_date = row['Date']

        for coad, sku, qty in zip(Coad, mapped_SKUs, quantities):
            print(po_number, Final_Address, coad, sku, qty, GSTN, company, start_date)

            sql = """
                INSERT INTO compiledpurchaseorder
                (ponumber, address, code, sku, quantity, gst, company, valuationdate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(ponumber, code)
                DO UPDATE SET
                    sku = excluded.sku,
                    address = excluded.address,
                    quantity = excluded.quantity,
                    gst = excluded.gst,
                    company = excluded.company,
                    valuationdate = excluded.valuationdate;
            """

            values = (
                po_number, Final_Address, coad, sku,
                qty, GSTN, company, start_date
            )

            cur.execute(sql, values)

    conn.commit()
    cur.close()
    conn.close()



